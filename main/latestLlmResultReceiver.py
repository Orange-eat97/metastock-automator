from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from phase2RequestReceiver import AddExplorerRequest, ExplorerColumn
from requestReceiver import parse_instruments_text


DEFAULT_LLM_RESULTS_PATH = Path(
    r"C:\GitHub\metastock-RAG-LLM\data\explorer_outputs.xlsx"
)

SHEET_NAME = "explorers"

REQUIRED_HEADERS = [
    "created_at",
    "backend",
    "model",
    "user_query",
    "validation_passed",
    "validation_errors_json",
    "full_output_json",
]


def _as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value

    if value is None:
        return False

    return str(value).strip().lower() in {"true", "1", "yes", "y", "passed"}


def _json_loads(value: Any, field_name: str) -> Any:
    if value is None:
        raise ValueError(f"'{field_name}' is empty.")

    if isinstance(value, (list, dict)):
        return value

    text = str(value).strip()

    if not text:
        raise ValueError(f"'{field_name}' is empty.")

    try:
        return json.loads(text)
    except json.JSONDecodeError as e:
        raise ValueError(
            f"Could not parse '{field_name}' as JSON.\n"
            f"Value was:\n{text}"
        ) from e


def _parse_columns(full_output: dict[str, Any]) -> list[ExplorerColumn]:
    raw_columns = full_output.get("col_definitions", [])

    if raw_columns is None:
        raw_columns = []

    if not isinstance(raw_columns, list):
        raise ValueError(
            "'full_output_json.col_definitions' must be a list."
        )

    columns: list[ExplorerColumn] = []

    for col in raw_columns:
        if not isinstance(col, dict):
            raise ValueError(f"Column definition must be a dict, got: {col!r}")

        slot = str(col.get("col_letter", "")).strip().upper()
        code_body = str(col.get("col_code", "")).strip()

        if not slot:
            raise ValueError(f"Column is missing 'col_letter': {col}")

        if not code_body:
            raise ValueError(f"Column {slot} is missing 'col_code': {col}")

        columns.append(
            ExplorerColumn(
                slot=slot,
                code_body=code_body,
            )
        )

    columns.sort(key=lambda c: ord(c.slot) - ord("A"))

    expected_slots = [chr(ord("A") + i) for i in range(len(columns))]
    actual_slots = [c.slot for c in columns]

    if actual_slots != expected_slots:
        raise ValueError(
            "Column slots must start from A and be continuous. "
            f"Expected {expected_slots}, got {actual_slots}."
        )

    return columns


def _read_rows_as_dicts(excel_path: Path) -> list[dict[str, Any]]:
    """
    Read the explorers sheet.

    Excel shape:
        row 1: headers
        row 2 onward: generated explorer rows

    Important:
        The generated Explorer object is read only from full_output_json.
    """
    if not excel_path.exists():
        raise FileNotFoundError(f"LLM results Excel not found: {excel_path}")

    wb = load_workbook(excel_path, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Sheet {SHEET_NAME!r} not found in {excel_path}. "
            f"Available sheets: {wb.sheetnames}"
        )

    ws = wb[SHEET_NAME]

    if ws.max_row < 2:
        raise ValueError(
            f"Sheet {SHEET_NAME!r} does not contain data rows. "
            f"Expected data from row 2 onward, got max_row={ws.max_row}."
        )

    headers = [
        str(ws.cell(row=1, column=col).value).strip()
        if ws.cell(row=1, column=col).value is not None
        else ""
        for col in range(1, ws.max_column + 1)
    ]

    missing = [h for h in REQUIRED_HEADERS if h not in headers]

    if missing:
        raise ValueError(
            "Excel header mismatch.\n"
            f"Missing headers: {missing}\n"
            f"Actual headers: {headers}"
        )

    result: list[dict[str, Any]] = []

    for row_index in range(2, ws.max_row + 1):
        item = {
            headers[col_index - 1]: ws.cell(row=row_index, column=col_index).value
            for col_index in range(1, ws.max_column + 1)
            if headers[col_index - 1]
        }

        # Skip completely empty rows.
        if not any(v is not None and str(v).strip() for v in item.values()):
            continue

        item["_excel_row"] = row_index
        result.append(item)

    return result


def _build_request_from_full_output(
    *,
    full_output: dict[str, Any],
    instruments_text: str | None,
    max_wait: int,
) -> AddExplorerRequest:
    name = str(full_output.get("explorer_name", "")).strip()
    notes = str(full_output.get("explorer_description", "")).strip()
    code_body = str(full_output.get("explorer_code_body", "")).strip()

    if not name:
        raise ValueError("'full_output_json.explorer_name' is empty.")

    if not code_body:
        raise ValueError("'full_output_json.explorer_code_body' is empty.")

    columns = _parse_columns(full_output)

    instrument_names, select_all = parse_instruments_text(instruments_text)

    if max_wait <= 0:
        raise ValueError("max_wait must be positive.")

    return AddExplorerRequest(
        # Phase 2 creation fields
        name=name,
        notes=notes,
        code_body=code_body,
        columns=columns,

        # Phase 1-compatible fields
        strategy_name=name,
        instrument_names=instrument_names,
        select_all_instruments=select_all,
        max_execution_wait_sec=max_wait,

        # Intended use for latest-LLM bridge.
        run_after_add=True,
    )


def load_latest_llm_add_explorer_request(
    *,
    excel_path: str | Path = DEFAULT_LLM_RESULTS_PATH,
    instruments_text: str | None = "all",
    max_wait: int = 300,
    require_validation_passed: bool = True,
) -> AddExplorerRequest:
    path = Path(excel_path)
    rows = _read_rows_as_dicts(path)

    if not rows:
        raise ValueError(f"No saved explorer rows found in {path}")

    latest_error: Exception | None = None

    for row in reversed(rows):
        if require_validation_passed and not _as_bool(row.get("validation_passed")):
            continue

        try:
            full_output = _json_loads(
                row.get("full_output_json"),
                field_name="full_output_json",
            )

            if not isinstance(full_output, dict):
                raise ValueError("'full_output_json' must contain a JSON object.")

            return _build_request_from_full_output(
                full_output=full_output,
                instruments_text=instruments_text,
                max_wait=max_wait,
            )

        except Exception as e:
            latest_error = e
            continue

    if latest_error:
        raise ValueError(
            "Could not build request from any usable latest row. "
            f"Last error: {latest_error}"
        ) from latest_error

    raise ValueError(
        "Could not find a usable latest explorer row. "
        "Check 'validation_passed' and 'full_output_json'."
    )


def describe_request(request: AddExplorerRequest) -> str:
    column_lines = [
        f"  col {col.slot} = {col.code_body}"
        for col in request.columns
    ]

    columns_text = "\n".join(column_lines) if column_lines else "  <none>"

    return (
        "\n=== Latest LLM Explorer Request ===\n"
        f"Name: {request.name}\n"
        f"Notes: {request.notes}\n"
        f"Filter code:\n{request.code_body}\n\n"
        f"Columns:\n{columns_text}\n\n"
        f"Select all instruments: {request.select_all_instruments}\n"
        f"Instrument names: {request.instrument_names}\n"
        f"Max wait: {request.max_execution_wait_sec}\n"
    )